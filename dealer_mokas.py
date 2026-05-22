import os
import re
from openpyxl import load_workbook

SRC_PATH = r"C:\EL\REPORT_SPPI\src\DB-OWNER-DEALER.xlsx"
DEST_DIR = r"C:\EL\REPORT_SPPI\asset"
DEST_PATH = os.path.join(DEST_DIR, "DB-OWNER-DEALER-MODIFIED.xlsx")
SHEET_NAME = "DBN"

PRE_TITLE_MAP: dict[str, str] = {
    "PROF": "Prof.",
    "DR": "Dr.",
    "IR": "Ir.",
    "DRS": "Drs.",
    "DRA": "Dra.",
    "H": "H.",
    "HJ": "Hj.",
    "HM": "H.M.",
    "NY": "Ny.",
    "TN": "Tn.",
    "RR": "Rr.",
    "MOH": "Moh.",
    "MUH": "Muh.",
}

POST_DEGREE_MAP: dict[str, str] = {
    "SE": "S.E",
    "SH": "S.H",
    "ST": "S.T",
    "SI": "S.I",
    "SIP": "S.I.P",
    "SSOS": "S.Sos",
    "SKOM": "S.Kom",
    "SAG": "S.Ag",
    "SPD": "S.Pd",
    "SKM": "S.K.M",
    "SFARM": "S.Farm",
    "SKEP": "S.Kep",
    "SGZ": "S.Gz",
    "SP": "S.P",
    "AMD": "A.Md",
    "MM": "M.M",
    "MT": "M.T",
    "MSI": "M.Si",
    "MPD": "M.Pd",
    "MH": "M.H",
    "MBA": "M.B.A",
    "PHD": "Ph.D",
    "MKES": "M.Kes",
    "MKOM": "M.Kom",
    "MSC": "M.Sc",
    "MAK": "M.Ak",
    "AK": "Ak",
    "CPA": "C.P.A",
    "CFP": "C.F.P",
    "CIA": "C.I.A",
    "NS": "Ns",
    "SAB": "S.A.B",
    "SHUT": "S.Hut",
    "SIKOM": "S.I.Kom",
    "SIK": "S.I.K",
    "SKED": "S.Ked",
    "MGZ": "M.Gz",
}

POST_DEGREE_PAIR_MAP: dict[tuple[str, str], str] = {
    ("S", "KOM"): "S.Kom",
    ("S", "SOS"): "S.Sos",
    ("S", "PD"): "S.Pd",
    ("S", "AG"): "S.Ag",
    ("S", "FARM"): "S.Farm",
    ("S", "KEP"): "S.Kep",
    ("S", "GZ"): "S.Gz",
    ("S", "AB"): "S.A.B",
    ("S", "HUT"): "S.Hut",
    ("S", "IKOM"): "S.I.Kom",
    ("S", "IK"): "S.I.K",
    ("S", "KED"): "S.Ked",
    ("M", "SI"): "M.Si",
    ("M", "PD"): "M.Pd",
    ("M", "KES"): "M.Kes",
    ("M", "KOM"): "M.Kom",
    ("M", "SC"): "M.Sc",
    ("M", "AK"): "M.Ak",
    ("M", "GZ"): "M.Gz",
    ("A", "MD"): "A.Md",
}

PARTICLES: frozenset[str] = frozenset({"BIN", "BINTI", "VAN", "DE", "AL", "EL"})

BALINESE_SECONDARY: frozenset[str] = frozenset(
    {
        "DEWA",
        "GUSTI",
        "KETUT",
        "MADE",
        "NENGAH",
        "WAYAN",
        "GEDE",
        "GDE",
        "KADEK",
        "KOMANG",
        "NYOMAN",
        "PUTU",
        "BAGUS",
        "ALIT",
        "LUH",
        "AGUNG",
        "COKORDA",
    }
)


def _key(token: str) -> str:
    return token.replace(".", "").upper()


def normalize_dealer_name(raw: str) -> str:
    if not isinstance(raw, str) or not raw.strip():
        return raw

    raw = raw.strip()
    raw = re.sub(r"\s+", " ", raw)
    raw = re.sub(r"\b(PT|CV|UD)\.(?=[a-zA-Z])", r"\1. ", raw, flags=re.IGNORECASE)
    raw = re.sub(r"\(\s+", "(", raw)
    raw = re.sub(r"\s+\)", ")", raw)
    raw = re.sub(r"\s*\(\s*", " (", raw)
    raw = raw.title()
    raw = re.sub(r"\bPt\b\.?", "PT", raw)
    raw = re.sub(r"\bCv\b\.?", "CV", raw)
    raw = re.sub(r"\bUd\b\.?", "UD", raw)
    return re.sub(r"\s+", " ", raw).strip()


def normalize_personal_name(raw: str) -> str:
    if not isinstance(raw, str) or not raw.strip():
        return raw

    raw = raw.strip()
    raw = raw.replace("..", ".")
    raw = re.sub(r"\b([a-zA-Z])\.(?=[a-zA-Z]{2,})", r"\1. ", raw)
    raw = re.sub(r",+$", "", raw.strip())
    raw = re.sub(r"\s*,\s*", ", ", raw)

    tokens: list[str] = [t for t in re.split(r"[\s,]+", raw) if t]
    if not tokens:
        return raw

    pre_titles: list[str] = []
    post_degrees: list[str] = []

    while tokens and _key(tokens[0]) in PRE_TITLE_MAP:
        pre_titles.append(PRE_TITLE_MAP[_key(tokens.pop(0))])

    while tokens:
        consumed = False
        if len(tokens) >= 2:
            pair = (_key(tokens[-2]), _key(tokens[-1]))
            if pair in POST_DEGREE_PAIR_MAP:
                post_degrees.insert(0, POST_DEGREE_PAIR_MAP[pair])
                tokens.pop()
                tokens.pop()
                consumed = True

        if not consumed and _key(tokens[-1]) in POST_DEGREE_MAP:
            post_degrees.insert(0, POST_DEGREE_MAP[_key(tokens.pop())])
            consumed = True

        if not consumed:
            break

    name_parts: list[str] = []
    for idx, token in enumerate(tokens):
        k = _key(token)

        if k in PRE_TITLE_MAP:
            name_parts.append(PRE_TITLE_MAP[k])
        elif k in PARTICLES:
            name_parts.append(token.lower())
        else:
            core = token.rstrip(".")
            if len(core) == 1 and core.isalpha():
                core = core.upper()

                if (
                    core == "I"
                    and (idx + 1 < len(tokens))
                    and _key(tokens[idx + 1]) in BALINESE_SECONDARY
                ):
                    name_parts.append("I")
                else:
                    name_parts.append(core + ".")
            else:
                name_parts.append(core.title())

    result = " ".join(pre_titles + name_parts)
    if post_degrees:
        result += ", " + ", ".join(post_degrees)

    return re.sub(r"\s+", " ", result).strip()


def _build_column_index(sheet, targets: list[str]) -> dict[str, int]:
    col_map: dict[str, int] = {}
    for cell in sheet[1]:
        header = str(cell.value).strip() if cell.value is not None else ""
        if header in targets:
            col_map[header] = cell.column

    missing = [c for c in targets if c not in col_map]
    if missing:
        print(
            f"[*] Peringatan: Kolom {missing} tidak ditemukan di sheet '{sheet.title}'. "
            f"Script tetap melanjutkan untuk memproses kolom yang tersedia."
        )

    return col_map


def process_dbn_sheet(sheet) -> None:
    targets = ["NAMA_DEALER", "NAMA_MITRA", "NAMA_BM", "NAMA_AM"]
    col_map = _build_column_index(sheet, targets)

    dealer_col = col_map.get("NAMA_DEALER")
    mitra_col = col_map.get("NAMA_MITRA")
    bm_col = col_map.get("NAMA_BM")
    am_col = col_map.get("NAMA_AM")

    for row in sheet.iter_rows(min_row=2):

        if dealer_col and row[dealer_col - 1].value is not None:
            row[dealer_col - 1].value = normalize_dealer_name(
                str(row[dealer_col - 1].value)
            )

        if mitra_col and row[mitra_col - 1].value is not None:
            row[mitra_col - 1].value = normalize_personal_name(
                str(row[mitra_col - 1].value)
            )

        if bm_col and row[bm_col - 1].value is not None:
            row[bm_col - 1].value = normalize_personal_name(str(row[bm_col - 1].value))

        if am_col and row[am_col - 1].value is not None:
            row[am_col - 1].value = normalize_personal_name(str(row[am_col - 1].value))


def main() -> None:
    if not os.path.isfile(SRC_PATH):
        raise FileNotFoundError(f"Source file not found: {SRC_PATH}")

    os.makedirs(DEST_DIR, exist_ok=True)

    print("[*] Memuat Workbook Excel...")
    wb = load_workbook(SRC_PATH)

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{SHEET_NAME}' tidak ditemukan. Sheet tersedia: {wb.sheetnames}"
        )

    print(f"[*] Memproses lembar sheet: {SHEET_NAME}")
    process_dbn_sheet(wb[SHEET_NAME])

    print(f"[*] Menyimpan Workbook modifikasi...")
    wb.save(DEST_PATH)
    print(f"[✔] Selesai. Hasil sukses ditulis ke:\n    {DEST_PATH}")


if __name__ == "__main__":
    main()
